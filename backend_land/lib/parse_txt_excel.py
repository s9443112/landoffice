from fpdf import FPDF
from os import listdir
from os.path import isfile, isdir, join, exists
import openpyxl
import datetime
import numpy as np
from copy import copy
from backend_land import models


def copy_sheet(source_sheet, target_sheet):
    # copy all the cel values and styles
    copy_cells(source_sheet, target_sheet)
    copy_sheet_attributes(source_sheet, target_sheet)


def copy_sheet_attributes(source_sheet, target_sheet):
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)
    # set row dimensions
    # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])
    if source_sheet.sheet_format.defaultColWidth is None:
        print('Unable to copy default column wide')
    else:
        target_sheet.sheet_format.defaultColWidth = copy(
            source_sheet.sheet_format.defaultColWidth)
    # set specific column width and hidden property
    # we cannot copy the entire column_dimensions attribute so we copy selected attributes
    for key, value in source_sheet.column_dimensions.items():
        # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
        target_sheet.column_dimensions[key].min = copy(
            source_sheet.column_dimensions[key].min)
        # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not onl;y the hidden property
        target_sheet.column_dimensions[key].max = copy(
            source_sheet.column_dimensions[key].max)
        target_sheet.column_dimensions[key].width = copy(
            source_sheet.column_dimensions[key].width)  # set width for every column
        target_sheet.column_dimensions[key].hidden = copy(
            source_sheet.column_dimensions[key].hidden)


def copy_cells(source_sheet, target_sheet):
    for (row, col), source_cell in source_sheet._cells.items():
        target_cell = target_sheet.cell(column=col, row=row)
        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)
        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)


class StartParseExcel():
    def __init__(self):
        self.ouput_path = "./output_land"
        self.fileName = ''

    def find_context(self, target):

        target = "["+target+"]"
        f = open(self.fileName, encoding="utf-8")
        text = f.read()
        f.close
        index = text.find(target) + len(target)
        text_ = text[index:]
        index2 = text_.find("[")
        if index2 == -1:
            return text[index+1:]
        return text[index+1:index2+index-1]

    def main(self):
        folders = listdir(self.ouput_path)
        workbook = openpyxl.load_workbook('tester_new2.xlsx')
        sheet_b = workbook['土地標示部']
        sheet2_b = workbook['土地所有權部']
        sheet3_b = workbook['土地權利部']
        sheet4_b = workbook['土地所有權部-歷次取得權利範圍']
       

        workbook2 = openpyxl.Workbook()
        workbook2.create_sheet('土地標示部')
        workbook2.create_sheet('土地所有權部')
        workbook2.create_sheet('土地權利部')
        workbook2.create_sheet('土地所有權部-歷次取得權利範圍')
       

        sheet = workbook2['土地標示部']
        sheet2 = workbook2['土地所有權部']
        sheet3 = workbook2['土地權利部']
        sheet4 = workbook2['土地所有權部-歷次取得權利範圍']
       
        try:
            copy_sheet(sheet_b, sheet)
            copy_sheet(sheet2_b, sheet2)
            copy_sheet(sheet3_b, sheet3)
            copy_sheet(sheet4_b, sheet4)
           
        except Exception as e:
            print(e)

        

        MarkingDepartments = models.MarkingDepartment.objects.all()
        OwnershipDepartments = models.OwnershipDepartment.objects.all()
        OtherShipDepartments = models.OtherShipDepartment.objects.all()
        Buildings_line = 2
        Publics_line = 2
        Location_line = 2

        for idx, each in enumerate(MarkingDepartments):
            sheet.cell(idx+2, 1).value = each.sheet1
            sheet.cell(idx+2, 2).value = each.sheet2
            sheet.cell(idx+2, 3).value = each.sheet3
            sheet.cell(idx+2, 4).value = each.sheet4
            sheet.cell(idx+2, 5).value = each.sheet5
            sheet.cell(idx+2, 6).value = each.sheet6
            sheet.cell(idx+2, 7).value = each.sheet7
            sheet.cell(idx+2, 8).value = each.sheet8
            sheet.cell(idx+2, 9).value = each.sheet9
            sheet.cell(idx+2, 10).value = each.sheet10
            sheet.cell(idx+2, 11).value = each.sheet11
            sheet.cell(idx+2, 12).value = each.sheet12
            sheet.cell(idx+2, 13).value = each.sheet13
            sheet.cell(idx+2, 14).value = each.sheet14
            sheet.cell(idx+2, 15).value = each.sheet15
            sheet.cell(idx+2, 16).value = each.sheet16

            

           

        for idx, each in enumerate(OwnershipDepartments):
            sheet2.cell(idx+2, 1).value = each.sheet1
            sheet2.cell(idx+2, 2).value = each.sheet2
            sheet2.cell(idx+2, 3).value = each.sheet3
            sheet2.cell(idx+2, 4).value = each.sheet4
            sheet2.cell(idx+2, 5).value = each.sheet5
            sheet2.cell(idx+2, 6).value = each.sheet6
            sheet2.cell(idx+2, 7).value = each.sheet7
            sheet2.cell(idx+2, 8).value = each.sheet8
            sheet2.cell(idx+2, 9).value = each.sheet9
            sheet2.cell(idx+2, 10).value = each.sheet10
            sheet2.cell(idx+2, 11).value = each.sheet11
            sheet2.cell(idx+2, 12).value = each.sheet12
            sheet2.cell(idx+2, 13).value = each.sheet13
            sheet2.cell(idx+2, 14).value = each.sheet14
            sheet2.cell(idx+2, 15).value = each.sheet15
            sheet2.cell(idx+2, 16).value = each.sheet16
            sheet2.cell(idx+2, 17).value = each.sheet17
            sheet2.cell(idx+2, 18).value = each.sheet18

            Buildings = models.OwnershipDepartmentHistory.objects.filter(ownershipdepartment=each)
          

            for idxx, eeach in enumerate(Buildings):
                sheet4.cell(Buildings_line, 1).value = each.sheet3+' / '+each.sheet4
                sheet4.cell(Buildings_line, 2).value = eeach.sheet1
                sheet4.cell(Buildings_line, 3).value = eeach.sheet2
                sheet4.cell(Buildings_line, 4).value = eeach.sheet3
                
                Buildings_line = Buildings_line+1

        for idx, each in enumerate(OtherShipDepartments):
            sheet3.cell(idx+2, 1).value = each.sheet1
            sheet3.cell(idx+2, 2).value = each.sheet2
            sheet3.cell(idx+2, 3).value = each.sheet3
            sheet3.cell(idx+2, 4).value = each.sheet4
            sheet3.cell(idx+2, 5).value = each.sheet5
            sheet3.cell(idx+2, 6).value = each.sheet6
            sheet3.cell(idx+2, 7).value = each.sheet7
            sheet3.cell(idx+2, 8).value = each.sheet8
            sheet3.cell(idx+2, 9).value = each.sheet9
            sheet3.cell(idx+2, 10).value = each.sheet10
            sheet3.cell(idx+2, 11).value = each.sheet11
            sheet3.cell(idx+2, 12).value = each.sheet12
            sheet3.cell(idx+2, 13).value = each.sheet13
            sheet3.cell(idx+2, 14).value = each.sheet14
            sheet3.cell(idx+2, 15).value = each.sheet15
            sheet3.cell(idx+2, 16).value = each.sheet16
            sheet3.cell(idx+2, 17).value = each.sheet17
            sheet3.cell(idx+2, 18).value = each.sheet18
            sheet3.cell(idx+2, 19).value = each.sheet19
            sheet3.cell(idx+2, 20).value = each.sheet20
            sheet3.cell(idx+2, 21).value = each.sheet21
            sheet3.cell(idx+2, 22).value = each.sheet22
            sheet3.cell(idx+2, 23).value = each.sheet23
            sheet3.cell(idx+2, 24).value = each.sheet24
            sheet3.cell(idx+2, 25).value = each.sheet25
            sheet3.cell(idx+2, 26).value = each.sheet26
            sheet3.cell(idx+2, 27).value = each.sheet27
            sheet3.cell(idx+2, 28).value = each.sheet28
            sheet3.cell(idx+2, 29).value = each.sheet29
        

        tonow = datetime.datetime.now()
        file_name = '{}/land_{}-{}-{}.xlsx'.format(self.ouput_path, tonow.year, tonow.month, tonow.day)
        count = 0
        while True:
            if exists(file_name) == True:
                count = count +1
                file_name = '{}/land_{}-{}-{}_{}.xlsx'.format(self.ouput_path, tonow.year, tonow.month, tonow.day,count)
            else:
                break
        workbook2.save(file_name)
        return file_name
            