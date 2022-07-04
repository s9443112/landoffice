from fpdf import FPDF
from os import listdir
from os.path import isfile, isdir, join, exists
import openpyxl
import datetime
import numpy as np
from copy import copy
from backend import models


def copy_sheet(source_sheet, target_sheet):
    # copy all the cel values and styles
    copy_cells(source_sheet, target_sheet)
    copy_sheet_attributes(source_sheet, target_sheet)


def copy_sheet_attributes(source_sheet, target_sheet):
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)
    # set row dimensions
    # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])
    if source_sheet.sheet_format.defaultColWidth is None:
        print('Unable to copy default column wide')
    else:
        target_sheet.sheet_format.defaultColWidth = copy(
            source_sheet.sheet_format.defaultColWidth)
    # set specific column width and hidden property
    # we cannot copy the entire column_dimensions attribute so we copy selected attributes
    for key, value in source_sheet.column_dimensions.items():
        # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
        target_sheet.column_dimensions[key].min = copy(
            source_sheet.column_dimensions[key].min)
        # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not onl;y the hidden property
        target_sheet.column_dimensions[key].max = copy(
            source_sheet.column_dimensions[key].max)
        target_sheet.column_dimensions[key].width = copy(
            source_sheet.column_dimensions[key].width)  # set width for every column
        target_sheet.column_dimensions[key].hidden = copy(
            source_sheet.column_dimensions[key].hidden)


def copy_cells(source_sheet, target_sheet):
    for (row, col), source_cell in source_sheet._cells.items():
        target_cell = target_sheet.cell(column=col, row=row)
        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)
        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)


class CusStartParseExcel():
    def __init__(self):
        self.ouput_path = "./output"
        self.fileName = ''

    def find_context(self, target):

        target = "["+target+"]"
        f = open(self.fileName, encoding="utf-8")
        text = f.read()
        f.close
        index = text.find(target) + len(target)
        text_ = text[index:]
        index2 = text_.find("[")
        if index2 == -1:
            return text[index+1:]
        return text[index+1:index2+index-1]

    def main(self, queryset: list =None):
        folders = listdir(self.ouput_path)
        workbook = openpyxl.load_workbook('tester_custom.xlsx')
        sheet_b = workbook['總輸出']
       
        workbook2 = openpyxl.Workbook()
        workbook2.create_sheet('總輸出')
       

        sheet = workbook2['總輸出']
       

        try:
            copy_sheet(sheet_b, sheet)
          
        except Exception as e:
            print(e)

        if queryset == None:
            MarkingDepartments = models.MarkingDepartment.objects.all()
        else:
            MarkingDepartments = queryset
       
        row_line = 2

        for idx, each in enumerate(MarkingDepartments):
            sheet.cell(row_line, 1).value = each.sheet1
            sheet.cell(row_line, 2).value = each.sheet2
            sheet.cell(row_line, 3).value = each.sheet3
            sheet.cell(row_line, 4).value = each.sheet4
            sheet.cell(row_line, 5).value = each.sheet7
            sheet.cell(row_line, 6).value = each.sheet8
            sheet.cell(row_line, 7).value = each.sheet10
            sheet.cell(row_line, 8).value = each.sheet11
            sheet.cell(row_line, 9).value = each.sheet12
            sheet.cell(row_line, 10).value = each.sheet13
            sheet.cell(row_line, 11).value = each.sheet14

            Buildings = models.MarkingDepartmentDependsBuildings.objects.filter(markingdepartment=each)
            Publics = models.MarkingDepartmentPublicPart.objects.filter(markingdepartment=each)
            Locations = models.MarkingDepartmentDependsLocationNumber.objects.filter(markingdepartment=each)

            output1 = ''
            for idxx, eeach in enumerate(Locations):
                output1 = output1 + eeach.sheet1 + ', '
                
            sheet.cell(row_line, 12).value = output1

            output1 = ''
            for idxx, eeach in enumerate(Buildings):
                output1 = output1 + eeach.sheet1 + ', ' + eeach.sheet2 + '\n'
            sheet.cell(row_line, 13).value = output1

            output1 = ''
            for idxx, eeach in enumerate(Publics):

                output1 = output1 + eeach.sheet2 + ', ' + eeach.sheet3 + '\n'
            sheet.cell(row_line, 14).value = output1


            OtherShipDepartments = models.OtherShipDepartment.objects.filter(markingdepartment=each)
            output1 = ''
            for idxx, eeach in enumerate(OtherShipDepartments):
                output1 = output1 + eeach.sheet9 + ', ' + eeach.sheet11+ ', ' + eeach.sheet15 + '\n'
            sheet.cell(row_line, 15).value = output1


            OwnershipDepartments = models.OwnershipDepartment.objects.filter(markingdepartment=each)

            for idxx, eeach in enumerate(OwnershipDepartments):
                sheet.cell(row_line, 16).value = eeach.sheet5
                sheet.cell(row_line, 17).value = eeach.sheet6
                sheet.cell(row_line, 18).value = eeach.sheet7
                sheet.cell(row_line, 19).value = eeach.sheet9
                sheet.cell(row_line, 20).value = eeach.sheet10
                sheet.cell(row_line, 21).value = eeach.sheet11
                sheet.cell(row_line, 22).value = eeach.sheet12

                row_line = row_line + 1
            
            
                

        tonow = datetime.datetime.now()
        file_name = '{}/custom_{}-{}-{}.xlsx'.format(self.ouput_path, tonow.year, tonow.month, tonow.day)
        count = 0
        while True:
            if exists(file_name) == True:
                count = count +1
                file_name = '{}/custom_{}-{}-{}_{}.xlsx'.format(self.ouput_path, tonow.year, tonow.month, tonow.day,count)
            else:
                break
        workbook2.save(file_name)
        return file_name
            