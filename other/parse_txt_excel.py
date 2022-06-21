from fpdf import FPDF
from os import listdir
from os.path import isfile, isdir, join, exists
import openpyxl
import datetime
import numpy as np
from copy import copy


def copy_sheet(source_sheet, target_sheet):
    # copy all the cel values and styles
    copy_cells(source_sheet, target_sheet)
    copy_sheet_attributes(source_sheet, target_sheet)


def copy_sheet_attributes(source_sheet, target_sheet):
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)
    # set row dimensions
    # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])
    if source_sheet.sheet_format.defaultColWidth is None:
        print('Unable to copy default column wide')
    else:
        target_sheet.sheet_format.defaultColWidth = copy(
            source_sheet.sheet_format.defaultColWidth)
    # set specific column width and hidden property
    # we cannot copy the entire column_dimensions attribute so we copy selected attributes
    for key, value in source_sheet.column_dimensions.items():
        # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
        target_sheet.column_dimensions[key].min = copy(
            source_sheet.column_dimensions[key].min)
        # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not onl;y the hidden property
        target_sheet.column_dimensions[key].max = copy(
            source_sheet.column_dimensions[key].max)
        target_sheet.column_dimensions[key].width = copy(
            source_sheet.column_dimensions[key].width)  # set width for every column
        target_sheet.column_dimensions[key].hidden = copy(
            source_sheet.column_dimensions[key].hidden)


def copy_cells(source_sheet, target_sheet):
    for (row, col), source_cell in source_sheet._cells.items():
        target_cell = target_sheet.cell(column=col, row=row)
        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)
        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)


class StartParseExcel():
    def __init__(self):
        self.ouput_path = "./output"
        self.fileName = ''

    def find_context(self, target):

        target = "["+target+"]"
        f = open(self.fileName, encoding="utf-8")
        text = f.read()
        f.close
        index = text.find(target) + len(target)
        text_ = text[index:]
        index2 = text_.find("[")
        if index2 == -1:
            return text[index+1:]
        return text[index+1:index2+index-1]

    def main(self):
        folders = listdir(self.ouput_path)
        workbook = openpyxl.load_workbook('tester.xlsx')
        sheet_b = workbook['標示部']
        sheet2_b = workbook['所有權部']
        sheet3_b = workbook['權利部']

        workbook2 = openpyxl.Workbook()
        workbook2.create_sheet('標示部')
        workbook2.create_sheet('所有權部')
        workbook2.create_sheet('權利部')

        sheet = workbook2['標示部']
        sheet2 = workbook2['所有權部']
        sheet3 = workbook2['權利部']

        try:
            copy_sheet(sheet_b, sheet)
            copy_sheet(sheet2_b, sheet2)
            copy_sheet(sheet3_b, sheet3)
        except Exception as e:
            print(e)

        start_row_sheet1 = 2
        start_row_sheet2 = 2
        start_row_sheet3 = 2
        for folder in folders:
            folder_fullpath = join(self.ouput_path, folder)
            if isdir(folder_fullpath):

                pdf = FPDF()
                pdf.add_font('font', "", "kaiu.ttf", True)
                pdf.set_font("font", size=12)

                files = listdir(folder_fullpath)
                for index, file in enumerate(files):

                    file_fullpath = join(folder_fullpath, file)

                    if isfile(file_fullpath):
                        print("檔案：", file)
                        if file == "標示部.txt":

                            print("觸發")
                            sheet.cell(start_row_sheet1, 1).value = "建物標示部"
                            pdf.add_page()
                            pdf.cell(200, 10, txt="＊＊＊　　建物標示部　　＊＊＊",
                                     ln=1, align="C")

                            self.fileName = file_fullpath

                            city = self.find_context("縣市")
                            town = self.find_context("鄉鎮市區")
                            section = self.find_context("地段")
                            buildnumber = self.find_context("建號")
                            pdf.cell(200, 10, txt=city+" " + town + " " +
                                     section+"   建號 "+buildnumber, ln=2, align="C")

                            sheet.cell(start_row_sheet1, 2).value = city
                            sheet.cell(start_row_sheet1, 3).value = town
                            sheet.cell(start_row_sheet1, 4).value = section
                            sheet.cell(start_row_sheet1, 5).value = buildnumber

                            date = self.find_context("登記日期")
                            reason = self.find_context("登記原因")
                            pdf.cell(200, 10, txt="登記日期:  "+date +
                                     "     登記原因：" + reason, ln=3, align="C")

                            sheet.cell(start_row_sheet1, 6).value = date
                            sheet.cell(start_row_sheet1, 7).value = reason

                            txt = self.find_context("建物門牌")
                            pdf.cell(200, 10, txt="建物門牌:  " +
                                     txt, ln=4, align="C")

                            sheet.cell(start_row_sheet1, 8).value = txt

                            txt = self.find_context("建物坐落地號")
                            pdf.cell(200, 10, txt="建物坐落地號" +
                                     " "*40, ln=5, align="C")
                            pdf.cell(200, 10, txt=txt, ln=6, align="C")
                            sheet.cell(start_row_sheet1, 9).value = '建物坐落地號'
                            sheet.cell(start_row_sheet1, 10).value = txt

                            usefor = self.find_context("主要用途")
                            element = self.find_context("主要建材")
                            pdf.cell(200, 10, txt="主要用途: "+usefor +
                                     "   主要建材: "+element, ln=7, align="C")

                            sheet.cell(start_row_sheet1, 11).value = usefor
                            sheet.cell(start_row_sheet1, 12).value = element

                            usefor = self.find_context("層數")
                            element = self.find_context("總面積")
                            pdf.cell(200, 10, txt="層數: "+usefor +
                                     "   總面積: "+element, ln=8, align="C")

                            sheet.cell(start_row_sheet1, 13).value = usefor

                            usefor = self.find_context("層次")
                            element = self.find_context("層次面積")
                            pdf.cell(200, 10, txt="層次: "+usefor +
                                     "   層次面積: "+element, ln=9, align="C")

                            sheet.cell(start_row_sheet1, 14).value = usefor

                            element = self.find_context("建築完成日期")
                            pdf.cell(200, 10, txt="建築完成日期: " +
                                     element, ln=10, align="C")

                            sheet.cell(start_row_sheet1, 15).value = element

                            ln = 11
                            cell = 17
                            #----------------------------------------------------------------------
                            sheet.cell(start_row_sheet1, 16).value = '附屬建物'

                            values = []
                            f = open(self.fileName, encoding="utf-8")
                            lines = f.readlines()
                            for line in lines:
                                values.append(line)
                            f.close
                            values = np.array(values)
                            print(values)
                            ii = np.where(values == '[附屬建物用途]\n')[0]

                            print(ii)
                            # input()

                            for each in ii:
                                # print(each)
                                pdf.cell(200, 10, txt="附屬建物用途: "+values[each+1] +"   面積: "+values[each+3], ln=ln, align="C")
                                sheet.cell(start_row_sheet1, cell).value = "附屬建物用途: "+values[each+1]+"   面積: "+values[each+3]
                                ln = ln +1
                                cell = cell +1
                            if len(ii) == 0:
                                cell  = cell +2
                            if len(ii) == 1:
                                cell  = cell +1
                            # usefor = self.find_context("附屬建物用途")
                            # element = self.find_context("面積")
                            # pdf.cell(200, 10, txt="附屬建物用途: "+usefor +"   面積: "+element, ln=ln, align="C")
                            # sheet.cell(start_row_sheet1, cell).value = "附屬建物用途: "+usefor+"   面積: "+element

                            #----------------------------------------------------------------------

                            buildnumber = self.find_context("其他登記事項")
                            pdf.cell(200, 10, txt="其他登記事項:", ln=ln, align="C")

                            sheet.cell(start_row_sheet1,cell).value = buildnumber

                            currentLine = ln +1
                            length = len(buildnumber)//30
                            for i in range(length):
                                pdf.cell(
                                    200, 10, txt=buildnumber[:30], ln=currentLine, align="C")
                                currentLine += 1
                                buildnumber = buildnumber[30:]
                            pdf.cell(200, 10, txt=buildnumber,
                                     ln=currentLine, align="C")

                            start_row_sheet1 = start_row_sheet1 + 1

                        elif file[:5] == "owner" and file[-3:] == "txt":

                            pdf.add_page()
                            pdf.cell(200, 10, txt="＊＊＊　　建物所有權部　　＊＊＊",
                                     ln=1, align="C")
                            # global fileName
                            self.fileName = file_fullpath

                            city = self.find_context("縣市名稱")
                            town = self.find_context("鄉鎮市區")
                            section = self.find_context("地段")
                            buildnumber = self.find_context("建號")
                            pdf.cell(200, 10, txt=city+" " + town + " " +
                                     section+"   建號 "+buildnumber, ln=2, align="C")
                            sheet2.cell(start_row_sheet2, 1).value = '建物所有權部'
                            sheet2.cell(start_row_sheet2, 2).value = city
                            sheet2.cell(start_row_sheet2, 3).value = town
                            sheet2.cell(start_row_sheet2, 4).value = section
                            sheet2.cell(start_row_sheet2,
                                        5).value = buildnumber

                            buildnumber = self.find_context("登記次序")
                            pdf.cell(200, 10, txt="登記次序:  " +
                                     buildnumber, ln=3, align="C")
                            sheet2.cell(start_row_sheet2,
                                        6).value = buildnumber

                            date = self.find_context("登記日期")
                            reason = self.find_context("登記原因")
                            pdf.cell(200, 10, txt="登記日期:  "+date +
                                     "     登記原因：" + reason, ln=4, align="C")
                            sheet2.cell(start_row_sheet2, 7).value = date
                            sheet2.cell(start_row_sheet2, 8).value = reason

                            buildnumber = self.find_context("原因發生日期")
                            pdf.cell(200, 10, txt="原因發生日期:  " +
                                     buildnumber, ln=5, align="C")
                            sheet2.cell(start_row_sheet2,
                                        9).value = buildnumber

                            buildnumber = self.find_context("所有權人姓名")
                            pdf.cell(200, 10, txt="所有權人姓名:  " +
                                     buildnumber, ln=6, align="C")
                            sheet2.cell(start_row_sheet2,
                                        11).value = buildnumber

                            buildnumber = self.find_context("統一編號")
                            pdf.cell(200, 10, txt="統一編號:  " +
                                     buildnumber, ln=7, align="C")
                            sheet2.cell(start_row_sheet2,
                                        12).value = buildnumber

                            buildnumber = self.find_context("住址")
                            pdf.cell(200, 10, txt="住址:  " +
                                     buildnumber, ln=8, align="C")
                            sheet2.cell(start_row_sheet2,
                                        13).value = buildnumber

                            buildnumber = self.find_context("權利範圍")
                            pdf.cell(200, 10, txt="權利範圍:  " +
                                     buildnumber, ln=9, align="C")
                            sheet2.cell(start_row_sheet2,
                                        15).value = buildnumber

                            buildnumber = self.find_context("權狀字號")
                            pdf.cell(200, 10, txt="權狀字號:  " +
                                     buildnumber, ln=10, align="C")
                            sheet2.cell(start_row_sheet2,
                                        16).value = buildnumber

                            buildnumber = self.find_context("其他登記事項")
                            pdf.cell(200, 10, txt="其他登記事項: ", ln=11, align="C")
                            sheet2.cell(start_row_sheet2,
                                        18).value = buildnumber

                            currentLine = 12
                            length = len(buildnumber)//30
                            for i in range(length):
                                pdf.cell(
                                    200, 10, txt=buildnumber[:30], ln=currentLine, align="C")
                                currentLine += 1
                                buildnumber = buildnumber[30:]
                            pdf.cell(200, 10, txt=buildnumber,
                                     ln=currentLine, align="C")
                            start_row_sheet2 = start_row_sheet2 + 1

                        elif file[:5] == "other" and file[-3:] == "txt":

                            pdf.add_page()
                            pdf.cell(200, 10, txt="＊＊＊　　建物他項權利部　　＊＊＊",
                                     ln=1, align="C")
                            # global fileName
                            self.fileName = file_fullpath

                            city = self.find_context("縣市名稱")
                            town = self.find_context("鄉鎮市區")
                            section = self.find_context("段號")
                            buildnumber = self.find_context("建號")
                            pdf.cell(200, 10, txt=city+" " + town + " " +
                                     section+"   建號 "+buildnumber, ln=2, align="C")

                            sheet3.cell(start_row_sheet3, 1).value = '建物他項權利部'
                            sheet3.cell(start_row_sheet3, 2).value = city
                            sheet3.cell(start_row_sheet3, 3).value = town
                            sheet3.cell(start_row_sheet3, 4).value = section
                            sheet3.cell(start_row_sheet3,
                                        5).value = buildnumber

                            buildnumber = self.find_context("登記次序")
                            buildnumber1 = self.find_context("權利種類")
                            pdf.cell(200, 10, txt="登記次序:  "+buildnumber +
                                     "   權利種類: "+buildnumber1, ln=3, align="C")
                            sheet3.cell(start_row_sheet3,
                                        6).value = buildnumber
                            sheet3.cell(start_row_sheet3,
                                        7).value = buildnumber1

                            date = self.find_context("收件年期")
                            reason = self.find_context("收件字號")
                            pdf.cell(200, 10, txt="收件年期:  "+date +
                                     "     收件字號：" + reason, ln=4, align="C")
                            sheet3.cell(start_row_sheet3, 8).value = date
                            sheet3.cell(start_row_sheet3, 9).value = reason

                            buildnumber = self.find_context("登記日期")
                            buildnumber1 = self.find_context("登記原因")
                            pdf.cell(200, 10, txt="登記日期:  "+buildnumber +
                                     "     登記原因: "+buildnumber1, ln=5, align="C")
                            sheet3.cell(start_row_sheet3,
                                        10).value = buildnumber
                            sheet3.cell(start_row_sheet3,
                                        11).value = buildnumber1

                            buildnumber = self.find_context("權利人姓名")
                            pdf.cell(200, 10, txt="權利人姓名:  " +
                                     buildnumber, ln=6, align="C")
                            sheet3.cell(start_row_sheet3,
                                        13).value = buildnumber

                            buildnumber = self.find_context("權利人統一編號")
                            pdf.cell(200, 10, txt="權利人統一編號:  " +
                                     buildnumber, ln=7, align="C")
                            sheet3.cell(start_row_sheet3,
                                        14).value = buildnumber

                            buildnumber = self.find_context("住址")
                            pdf.cell(200, 10, txt="住址:  " +
                                     buildnumber, ln=8, align="C")

                            buildnumber = self.find_context("債權額比例")
                            buildnumber1 = self.find_context("擔保債權總金額")
                            pdf.cell(200, 10, txt="債權額比例:  "+buildnumber +
                                     "     擔保債權總金額: "+buildnumber1, ln=9, align="C")
                            sheet3.cell(start_row_sheet3,
                                        17).value = buildnumber
                            sheet3.cell(start_row_sheet3,
                                        18).value = buildnumber1

                            pdf.cell(200, 10, txt="擔保債權種類及範圍:  ",
                                     ln=10, align="C")
                            currentLine = 11
                            buildnumber = self.find_context("擔保債權種類及範圍")
                            sheet3.cell(start_row_sheet3,
                                        19).value = buildnumber

                            length = len(buildnumber)//30
                            for i in range(length):
                                pdf.cell(
                                    200, 10, txt=buildnumber[:30], ln=currentLine, align="C")
                                currentLine += 1
                                buildnumber = buildnumber[30:]
                            pdf.cell(200, 10, txt=buildnumber,
                                     ln=currentLine, align="C")
                            currentLine += 1

                            buildnumber = self.find_context("擔保債權確定期日")
                            buildnumber1 = self.find_context("清償日期")
                            pdf.cell(200, 10, txt="擔保債權確定期日:  "+buildnumber +
                                     "    清償日期: "+buildnumber1, ln=currentLine, align="C")
                            currentLine += 1

                            sheet3.cell(start_row_sheet3,
                                        20).value = buildnumber
                            sheet3.cell(start_row_sheet3,
                                        21).value = buildnumber1

                            buildnumber = self.find_context("利息(率)或地租")
                            buildnumber1 = self.find_context("違約金")
                            pdf.cell(200, 10, txt="利息(率)或地租:  "+buildnumber +
                                     "    違約金: "+buildnumber1, ln=currentLine, align="C")
                            currentLine += 1

                            sheet3.cell(start_row_sheet3,
                                        22).value = buildnumber
                            sheet3.cell(start_row_sheet3,
                                        23).value = buildnumber1

                            element = self.find_context("其他擔保範圍約定")
                            lines = element.split("。")
                            pdf.cell(200, 10, txt="其他擔保範圍約定: ",
                                     ln=currentLine, align="C")

                            currentLine += 1
                            for index, line in enumerate(lines):
                                pdf.cell(200, 10, txt=line,
                                         ln=currentLine+index, align="C")
                            # sheet3.cell(start_row_sheet3,23).value = buildnumber1
                            start_row_sheet3 = start_row_sheet3 + 1
                pdf.output(join(folder_fullpath, 'result.pdf'), 'F')

        tonow = datetime.datetime.now()
        file_name = '{}/{}-{}-{}.xlsx'.format(self.ouput_path, tonow.year, tonow.month, tonow.day)
        count = 0
        while True:
            if exists(file_name) == True:
                count = count +1
                file_name = '{}/{}-{}-{}_{}.xlsx'.format(self.ouput_path, tonow.year, tonow.month, tonow.day,count)
            else:
                break
        workbook2.save(file_name)
